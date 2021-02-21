# -*- coding: utf-8 -*-
"""
Filename: hsb_trade_tracker.py
Date created: Fri Aug 21 12:17:57 2020
@author: Julio Hong
Purpose: Tracks all the ongoing trades in the bazaar of Hypixel Skyblock. Focus on lapis lazuli for now.
        Track how the amount per price changes over time for buy/sell orders.
        Maybe also track instant buy/sell
Steps: 
Find the lapis_info
Create a df for each price. Then each row represents a timestamp, and the data represents the amount transacted.
Create a dict of dfs? But updating them would be annoying. Maybe multi-index? Timestamp -> prices
But if a price is absent? Still have to add a timestamp row to show zero value.
It'll be wasteful I guess, but if timestamp -> price, some prices will be absent within timestamps.
Price -> Timestamp. Timestamps are columnns because easier to add to the df
If there is a new price, then add a new row to the outer index
Store the lapis_info in an external document, maybe a spreadsheet.
Refresh the scraping of lapis_info
Change to melons because lapis is full of speculation and price depression.
Graph the info

Notes: https://hypixel.net/threads/skyblock-cant-find-what-item-is-that.3064833/
Cowtipper said:
Hypixel SkyBlock uses the pre-Minecraft 1.13 damage/data values for some items: 
INK_SACK:0 is the normal ink sack (= a dye with damage value zero).
INK_SACK:3 is a dye with damage value 3: Cocoa Beans.
INK_SACK:4 is Lapis Lazuli respectively.

You can see the full list of all item>damage values on the Minecraft Wiki.
It works the same way for logs and fish: for example RAW_FISH:0 is the normal fish, while RAW_FISH:3 is pufferfish.
Check for the largest spread for bazaar flipping?
"""

import requests
import datetime as dt
from time import sleep
import pandas as pd
import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt
plt.rcParams["figure.figsize"] = [10, 5]

# To adjust the dataframe appearance
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 20)
pd.set_option('display.width', 200)


# sold_over_time_df = pd.DataFrame(index=)


# Unique to each player and can be updated
api_key = '4aa4f936-c170-4a8e-849a-4418f76c8fe5'
payload = {'key': api_key}
sold_over_time_dfs = {}
bought_over_time_dfs = {}
spread_over_time_df = pd.DataFrame([])
# Add the selected keys into a list
# goods_of_interest = ['MELON', 'IRON_INGOT', 'DIAMOND', 'WHEAT', 'OBSIDIAN']
# goods_of_interest = ['DIAMOND', 'INK_SACK:4']
goods_of_interest = 'ALL GOODS'


def export_create_or_update(filename, input_df, good=''):
    # Export the results
    # Use goods_of_interest to generate filenames?
    # Open the existing file and add a new sheet based on datetime
    # Might convert this into a function in future
    if good == '':
        # Check if the files already exist
        if os.path.exists(filename):
            print('Update existing spreadsheet')
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        else:
            print('Create new spreadsheet')
            writer = pd.ExcelWriter(filename, engine='xlsxwriter')

    else:
        # Determine if buy or sell sheet is being updated
        if 'sold' in filename:
            txn_type = 'sell'
        elif 'bought' in filename:
            txn_type = 'buy'
        # Check if the files already exist
        if os.path.exists(filename):
            print('Update existing ' + txn_type + '_spreadsheet for ' + good)
            # sold_over_time
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        else:
            print('Create new ' + txn_type + '_spreadsheet for ' + good)
            writer = pd.ExcelWriter(filename, engine='xlsxwriter')

    # Put all the common code after the 'else' statement
    input_df.to_excel(writer, sheet_name=timestamp)
    writer.save()
    writer.close()


def find_best_spread(bought_over_time_dfs, spread_over_time_df, interval=50):
    # Need to track number of transactions for each item also
    # See how popular they are and the spread as well or else I try to flip slow-moving items

    # If the delta is non-zero for a certain period of time, then mark it on the spread_df
    # This is kinda nasty. Just take the average within the latest interval and see if there are any transactions
    # See what is the best interval value to select? Run a few values

    # Then calculate a weight based on number of transactions within the interval. Price doesn't matter just the number of orders.
    # I removed the orders up earlier. Uh oh. I guess it's not that needed.

    def calc_timestamp_deltas(bought_over_time_dfs, interval):
        trade_freq = pd.Series(index=bought_over_time_dfs.keys())
        for item in bought_over_time_dfs.keys():
            # Track the delta between columns.
            deltas_over_time = bought_over_time_dfs[item].diff(axis=1)
            delta_average = deltas_over_time.sum().sum() / interval
            trade_freq.loc[item] = abs(delta_average)
        return trade_freq

    # I was going to take the latest value but average is less susceptible to random error
    latest_spreads = pd.DataFrame(index=spread_over_time_df.index)
    latest_spreads['mean_spread'] = spread_over_time_df.iloc[:,-interval:].mean(axis=1)
    # latest_spreads = spread_over_time_df.iloc[:,-interval:].mean(axis=1)
    latest_spreads['trade_freq_multiplier'] = calc_timestamp_deltas(bought_over_time_dfs, interval)
    # latest_spreads['buy_trades'] = calc_timestamp_deltas(bought_over_time_dfs, interval)
    # latest_spreads['sell_trades'] = calc_timestamp_deltas(sold_over_time_dfs, interval)
    # Take the min of the buying trades and selling trades, which might represent what is being satisfied?
    # latest_spreads['trade_freq_multiplier'] = latest_spreads[['buy_trades', 'sell_trades']].max(axis=1)
    # No, how to distinguish between trades being satisfied and trades being removed?

    # Find the sellPrice (bazaar sells)
    # Brutal way of doing this
    # for key in json_products.keys():
    #     # I'm not sure how this value is calculated, but it sometimes doesn't match the bazaar sell values
    #     latest_spreads.loc[key, 'sellPrice'] = json_products[key]['quick_status']['sellPrice']

    for key in sold_over_time_dfs.keys():
        # Find the lowest sell price within the interval
        # print(key)
        latest_means = sold_over_time_dfs[key].iloc[:, -interval:].mean(axis=1)
        # print(latest_means)
        non_nan_means = latest_means.dropna()
        try:
            latest_spreads.loc[key, 'sellPrice'] = max(list(non_nan_means.index))
        except ValueError:
            # Make this very high so it won't show up under best options
            latest_spreads.loc[key, 'sellPrice'] = -1
    # Calculate profit% from the spread
    latest_spreads['profit_pct'] = latest_spreads['mean_spread'] / latest_spreads['sellPrice']
    # Multiply it by the spread.
    # latest_spreads['weighted_spread'] = latest_spreads['mean_spread'] * latest_spreads['trade_freq_multiplier']
    latest_spreads['weighted_spread'] = latest_spreads['mean_spread'] * latest_spreads['trade_freq_multiplier'] * latest_spreads['profit_pct']
    # Suppress scientific notation in pandas
    # latest_spreads.apply(lambda x: '%.5f' % x)
    latest_spreads = latest_spreads.round(1)
    return latest_spreads


def load_goods_records(path):
    # Load the results and stitch them together?
    # Find all the files that exist in the directory
    recording_spreadsheets = [file for file in os.listdir(path) if file.endswith(".xlsx")]
    for file in recording_spreadsheets:
        filename_list = file.split('_')
        good = filename_list[0]
        # Combines all the sheets into a single dataframe
        read_df = pd.concat(pd.read_excel(file, sheet_name=None, index_col=0), axis=0)
        # Removes the highest level of the multi-index columns
        read_df.columns = read_df.columns.droplevel()
        # Sort by good

        # Sort sell or buy
        if filename_list[1] == 'sold':
            sold_over_time_dfs[good] = read_df
        elif filename_list[1] == 'bought':
            bought_over_time_dfs[good] = read_df
        else:
            return 'How did we get here?'


#=========== RUN CODE ===========

while True:
    try:
        start_time = dt.datetime.now()
        runtime_elapsed = dt.timedelta(minutes=1)
        runtime_cap = dt.timedelta(minutes=15)
        # Run this in a loop every interval
        # # 1) Testing mode
        # count = 0
        # while count < 4:
        #     count += 1

        # 2) Run mode
        # while True:
        # Run a while loop for every interval of 15min.
        while runtime_elapsed < runtime_cap:
            try:
                time_of_scrape = dt.datetime.now()

                # print('Hypixel Skyblock Bazaar trading for ' + str(goods_of_interest) + ' running at ' + str(time_of_scrape))
                print('Hypixel Skyblock Bazaar trading for ALL GOODS running at ' + str(time_of_scrape))
                r = requests.get('https://api.hypixel.net/skyblock/bazaar', params=payload)
                #r = requests.get('https://api.hypixel.net/skyblock/product?', params=payload)
                json_data = (r.json())
                json_products = json_data['products']

                # Get everything inside
                goods_of_interest = list(json_products.keys())
                sell_dfs = {}
                buy_dfs = {}

                # I'm going to generalise this. I don't know if that's wise
                for good in json_products.keys():
                    # Convert sell_summary and buy_summary into df
                    # Sell and buy are reversed for some reason
                    sell_dfs[good] = pd.DataFrame(json_products[good]['sell_summary'])
                    buy_dfs[good] = pd.DataFrame(json_products[good]['buy_summary'])

                    # If there are no transactions available at the time
                    if sell_dfs[good].empty or buy_dfs[good].empty:
                        print(good + ' has no transactions')
                        sell_dfs[good]['orders'] = -1
                        sell_dfs[good][time_of_scrape] = -1
                        buy_dfs[good]['orders'] = -1
                        buy_dfs[good][time_of_scrape] = -1

                        spread_over_time_df.loc[good, time_of_scrape] = -1

                    else:
                        # Change price to index
                        sell_dfs[good].set_index('pricePerUnit', inplace=True)
                        buy_dfs[good].set_index('pricePerUnit', inplace=True)

                        # Change amount to timestamp
                        sell_dfs[good].rename(columns={'amount':time_of_scrape}, inplace=True)
                        buy_dfs[good].rename(columns={'amount':time_of_scrape}, inplace=True)

                        # print(good + ' being scraped')
                        # Set a threshold volume
                        threshold_vol = 2000
                        # Find the highest buy price and lowest sell price above the threshold
                        highest_sell = max(sell_dfs[good].index)
                        lowest_buy = min(buy_dfs[good].index)

                        # Find delta and store in a df for each good against time
                        spread_over_time_df.loc[good, time_of_scrape] = lowest_buy - highest_sell

                    # If first, then set the main df as equal to this timestamp
                    # Check if this variable exists first, then concat with the new column
                    # print(("sold_over_time_dfs['" + good + "']"))
                    # if ("sold_over_time_dfs['" + good + "']") in locals():
                    # Change to check if the value exists in keys()
                    # Only if this is a GOI, then record in a df
                    if good in goods_of_interest:
                        if good in sold_over_time_dfs.keys():
                            sold_over_time_dfs[good] = pd.concat([sold_over_time_dfs[good], sell_dfs[good][time_of_scrape]], axis=1)
                            bought_over_time_dfs[good] = pd.concat([bought_over_time_dfs[good], buy_dfs[good][time_of_scrape]], axis=1)

                        # Else set the current df as the template
                        else:
                            print('Initialising over_time_dfs for ' + good)
                            sold_over_time_dfs[good] = sell_dfs[good].copy()
                            bought_over_time_dfs[good] = buy_dfs[good].copy()

                            # Temporary until I'm sure the amount is being tabulated properly
                            sold_over_time_dfs[good].drop(columns='orders', inplace=True)
                            bought_over_time_dfs[good].drop(columns='orders', inplace=True)

                print('Scrape completed at ' + str(dt.datetime.now()))
                sleep(10)
                # Measure how long the loop has been running
                runtime_elapsed = dt.datetime.now() - start_time

            # Adds a way to break the loop with user input
            except KeyboardInterrupt:
                break

        # Moved out of the except clause because I want this to also run after the while-loop ends
        # Put this here so I don't keep making new sheets every time I run the export section
        # Excel can't accept ":" in sheet name
        timestamp = str(dt.datetime.now()).replace(':', '_')
        print('End scraping')

        # Graph the results
        # Can't really make sense of it though...
        # bought_over_time_dfs['DIAMOND'].T.plot()

        # Save data
        folder_path = r"C:\Users\Julio Hong\Documents\GitHub\Minecraft\Hypixel Skyblock\\"
        print('Exporting results to Excel')
        sell_spreadsheets = {}
        buy_spreadsheets = {}
        goods_of_interest = ['DIAMOND', 'INK_SACK:4']
        for good in goods_of_interest:
            if ':' in good:
                good_name = good.replace(':', '_')
            else:
                good_name = good
            sell_spreadsheets[good] = os.path.join(folder_path, good_name + "_sold_over_time.xlsx")
            buy_spreadsheets[good] = os.path.join(folder_path, good_name + "_bought_over_time.xlsx")

            export_create_or_update(sell_spreadsheets[good], sold_over_time_dfs[good], good)
            export_create_or_update(buy_spreadsheets[good], bought_over_time_dfs[good], good)

        # If the loop time is less than the interval time OR If before the halfway mark, revert back to last spread_data
        # Elif after the halfway mark, calculate a new truncated spread data (and mark it)
        # if runtime_elapsed < dt.timedelta(seconds=interval * 15) / 2:
        interval = 50
        if runtime_elapsed < dt.timedelta(seconds=interval * 15):
            interval = len(spread_over_time_df.columns)

        more_spreads = find_best_spread(bought_over_time_dfs, spread_over_time_df, interval)
        print(more_spreads.sort_values(by=['weighted_spread']))

        # Export the spreads
        txn_spread_file = os.path.join(folder_path, "txn_spread_over_time.xlsx")
        export_create_or_update(txn_spread_file, more_spreads.sort_values(by=['weighted_spread'], ascending=False))
        print('Results successfully exported')

    except KeyboardInterrupt:
        interval = 50
        # If the loop time is less than the interval time OR If before the halfway mark, revert back to last spread_data
        # Elif after the halfway mark, calculate a new truncated spread data (and mark it)
        # if runtime_elapsed < dt.timedelta(seconds=interval * 15) / 2:
        if runtime_elapsed < dt.timedelta(seconds=interval * 15):
            interval = len(spread_over_time_df.columns)


        more_spreads = find_best_spread(bought_over_time_dfs, spread_over_time_df, interval)
        print(more_spreads.sort_values(by=['weighted_spread']))

        # Export the spreads
        txn_spread_file = os.path.join(folder_path, "txn_spread_over_time.xlsx")
        export_create_or_update(txn_spread_file, more_spreads.sort_values(by=['weighted_spread'], ascending=False))
        print('Results successfully exported')
        break

# Run this entire code until I want it to stop
# Loop it every 10-15 min

print('BEFORE SHARING REMOVE YOUR API KEY')













